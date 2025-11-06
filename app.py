import os, re, json, time, base64, mimetypes, tempfile, traceback
from uuid import uuid4
from flask import Flask, request, jsonify, send_from_directory, Response
import requests

# 追加: EML生成用
from email.message import EmailMessage
from email.utils import formatdate
import html as _html

# ===== excel_api.py 由来の追加 import（仕様変更なしで合体） =====
import html  # excel_api.py で使用
from io import BytesIO
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook
import xlrd  # .xls対応
from email import policy
from email.parser import BytesParser
import extract_msg  # .msg対応（excel_api 由来機能で使用）

app = Flask(__name__, static_folder="static", template_folder="templates")

# ===============================
# 設定
# ===============================
TENANT_ID         = os.getenv("AZ_TENANT_ID", "")
CLIENT_ID_ORG     = os.getenv("AZ_CLIENT_ID", "")
CLIENT_SECRET_ORG = os.getenv("AZ_CLIENT_SECRET", "")
TARGET_USER_ID    = os.getenv("TARGET_USER_ID", "")  # /users/{id} or UPN

GRAPH_BASE  = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPE = "https://graph.microsoft.com/.default"

SMALL_MAX_BYTES    = 250 * 1024 * 1024  # 250MB
CHUNK_SIZE         = 5   * 1024 * 1024  # 5MB per chunk
DEFAULT_TICKET_TTL = 600                # 10 min

# ===============================
# チケット & トークン
# ===============================
TICKETS = {}  # { ticket_id: { type, fileName, mime, data|payload|href, expire, ... } }
TOKENS  = {"access_token": "", "expire": 0}

def save_ticket(meta, ttl=DEFAULT_TICKET_TTL):
    tid = uuid4().hex
    meta["expire"] = time.time() + ttl
    TICKETS[tid] = meta
    return tid

def redeem_ticket(ticket, consume=True):
    if ticket not in TICKETS:
        raise KeyError("ticket_not_found_or_expired")
    meta = TICKETS[ticket]
    if meta.get("expire", 0) < time.time():
        raise KeyError("ticket_expired")
    if consume:
        TICKETS.pop(ticket, None)
    return meta

# ===============================
# Graph ヘルパ
# ===============================
def refresh_if_needed():
    if TOKENS["expire"] > time.time() + 60:
        return
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    r = requests.post(url, data={
        "client_id": CLIENT_ID_ORG,
        "client_secret": CLIENT_SECRET_ORG,
        "scope": GRAPH_SCOPE,
        "grant_type": "client_credentials"
    }, timeout=30)
    r.raise_for_status()
    j = r.json()
    TOKENS["access_token"] = j["access_token"]
    TOKENS["expire"] = time.time() + int(j.get("expires_in", 3600))

def _auth_headers(extra=None):
    refresh_if_needed()
    h = {"Authorization": f"Bearer {TOKENS['access_token']}"}
    if extra:
        h.update(extra)
    return h

def _sanitize_name(name: str) -> str:
    return (re.sub(r'[\\/:*?"<>|]', "_", name or "").strip() or "NewItem")

# ---- アップロード関連
def graph_put_small_to_folder_org(folder_id, name, mime, data):
    safe_name = _sanitize_name(name)
    url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{folder_id}:/{safe_name}:/content"
    return requests.put(url, headers=_auth_headers({"Content-Type": mime or "application/octet-stream"}),
                        data=data, timeout=300)

def graph_create_upload_session_org(folder_id, name):
    safe_name = _sanitize_name(name)
    url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{folder_id}:/{safe_name}:/createUploadSession"
    r = requests.post(url, headers=_auth_headers(), json={}, timeout=60)
    r.raise_for_status()
    return r.json()["uploadUrl"]

def graph_put_chunked_to_folder_org(folder_id, name, data):
    upload_url = graph_create_upload_session_org(folder_id, name)
    size = len(data); off = 0; last = None
    while off < size:
        chunk = data[off: off + CHUNK_SIZE]
        start = off; end = off + len(chunk) - 1
        headers = {
            "Content-Length": str(len(chunk)),
            "Content-Range": f"bytes {start}-{end}/{size}",
            "Content-Type": "application/octet-stream",
        }
        last = requests.put(upload_url, headers=headers, data=chunk, timeout=600)
        if last.status_code not in (200, 201, 202):
            break
        off += len(chunk)
    return last

# ---- メタ/フォルダ列挙/パンくず/作成/削除
def graph_get_item_meta(item_id):
    url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{item_id}"
    r = requests.get(url, headers=_auth_headers(), timeout=30)
    r.raise_for_status()
    return r.json()

def graph_list_child_folders(parent_id: str):
    """
    OneDriveの指定フォルダ配下のフォルダ一覧をすべて取得する
    （@odata.nextLink を追跡して全ページ取得）
    """
    if not parent_id or parent_id == "root":
        url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/root/children?$select=id,name,folder&$top=999"
    else:
        url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{parent_id}/children?$select=id,name,folder&$top=999"

    arr = []

    while url:
        r = requests.get(url, headers=_auth_headers(), timeout=30)
        r.raise_for_status()
        data = r.json()

        for it in data.get("value", []):
            if isinstance(it.get("folder"), dict):
                arr.append({"id": it.get("id"), "name": it.get("name")})

        url = data.get("@odata.nextLink")

    return arr

def graph_get_item_parent(item_id: str):
    url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{item_id}?$select=id,name,parentReference"
    r = requests.get(url, headers=_auth_headers(), timeout=30)
    r.raise_for_status()
    j = r.json()
    parent_id = (j.get("parentReference") or {}).get("id")
    return {"id": j.get("id"), "name": j.get("name"), "parentId": parent_id}

def graph_create_folder(parent_id: str, name: str, conflict_behavior="rename"):
    safe_name = _sanitize_name(name)
    if not parent_id or parent_id == "root":
        url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/root/children"
    else:
        url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{parent_id}/children"
    body = {
        "name": safe_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": conflict_behavior
    }
    r = requests.post(url, headers=_auth_headers({"Content-Type":"application/json"}), json=body, timeout=30)
    r.raise_for_status()
    return r.json()

def graph_delete_item(item_id: str):
    if not item_id:
        raise ValueError("missing item_id")
    url = f"{GRAPH_BASE}/users/{TARGET_USER_ID}/drive/items/{item_id}"
    r = requests.delete(url, headers=_auth_headers(), timeout=30)
    r.raise_for_status()
    return True

# ===============================
# EML生成（空EML対策の修正ポイント）
# ===============================
def _text_to_html(s: str) -> str:
    if not s:
        return "<p></p>"
    t = (_html.escape(s or "").replace("\r\n", "\n").replace("\r", "\n"))
    paras = t.split("\n\n")
    return "".join(f"<p>{p.replace('\n','<br>')}</p>" for p in paras) or "<p></p>"

def build_eml_bytes(subject, from_addr, to_addrs, body_text="", body_html=None, date_str=None) -> bytes:
    msg = EmailMessage()
    msg["Subject"] = subject or "LLM Output"
    msg["From"] = from_addr or "noreply@example.com"
    msg["To"] = ", ".join(to_addrs) if isinstance(to_addrs, list) else (to_addrs or "")
    msg["Date"] = date_str or formatdate(localtime=True)
    msg.set_content(body_text or "", subtype="plain", charset="utf-8")
    if body_html is None or body_html is False:
        body_html = _text_to_html(body_text or "")
    if body_html:
        msg.add_alternative(body_html, subtype="html", charset="utf-8")
    return msg.as_bytes()

# ===============================
# バイト展開 / .msg抽出 / .eml抽出（+ 任意添付抽出を追加）
# ===============================
def materialize_bytes(meta):
    """
    チケットメタから (file_name, bytes, mime) を返す
    type: "base64" / "text" / "eml" / "url"
    """
    t = (meta.get("type") or "").lower()
    if t == "base64":
        return meta.get("fileName") or "download.bin", base64.b64decode(meta.get("data") or ""), meta.get("mime") or "application/octet-stream"
    elif t == "text":
        return meta.get("fileName") or "note.txt", (meta.get("data") or "").encode("utf-8"), meta.get("mime") or "text/plain"
    elif t == "url":
        u = meta.get("href")
        if not u:
            raise RuntimeError("payload.url missing")
        r = requests.get(u, timeout=60)
        r.raise_for_status()
        return meta.get("fileName") or "download.bin", r.content, meta.get("mime") or "application/octet-stream"
    elif t == "eml":
        # 1) data（base64）があればそれを使う
        data_b64 = meta.get("data")
        if data_b64:
            return meta.get("fileName") or "message.eml", base64.b64decode(data_b64), meta.get("mime") or "message/rfc822"
        # 2) payload から EML を組み立てる
        p = meta.get("payload") or {}
        body_text = p.get("text") or p.get("body") or ""
        body_html = p.get("html")
        if body_html in (None, False, "") and (p.get("htmlFromText") or True):
            body_html = _text_to_html(body_text)
        eml = build_eml_bytes(
            subject=p.get("subject"),
            from_addr=p.get("from"),
            to_addrs=p.get("to") or ["user@example.com"],
            body_text=body_text,
            body_html=body_html,
            date_str=p.get("date"),
        )
        return meta.get("fileName") or "message.eml", eml, meta.get("mime") or "message/rfc822"
    else:
        return meta.get("fileName") or "download.bin", base64.b64decode(meta.get("data") or ""), meta.get("mime") or "application/octet-stream"

def _is_excel_filename(name: str) -> bool:
    n = (name or "").lower()
    return n.endswith((".xlsx", ".xlsm", ".xls", ".csv"))

def _is_excel_mime(mime: str) -> bool:
    m = (mime or "").lower()
    return (
        m.startswith("application/vnd.openxmlformats-officedocument.spreadsheetml")
        or m == "application/vnd.ms-excel"
        or m == "text/csv"
    )

def _is_pdf_filename(name: str) -> bool:
    return (name or "").lower().endswith(".pdf")

def _is_pdf_mime(mime: str) -> bool:
    return (mime or "").lower() == "application/pdf"

def _is_word_filename(name: str) -> bool:
    n = (name or "").lower()
    return n.endswith((".docx", ".doc"))

def _is_word_mime(mime: str) -> bool:
    m = (mime or "").lower()
    return m in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword")

def _classify_by_name_mime(name: Optional[str], mime: Optional[str]) -> Optional[str]:
    """ return 'excel' | 'pdf' | 'word' | None """
    if _is_excel_filename(name or "") or _is_excel_mime(mime or ""):
        return "excel"
    if _is_pdf_filename(name or "") or _is_pdf_mime(mime or ""):
        return "pdf"
    if _is_word_filename(name or "") or _is_word_mime(mime or ""):
        return "word"
    return None

def _extract_first_excel_from_msg(msg_bytes: bytes):
    import extract_msg  # pip install extract-msg
    with tempfile.NamedTemporaryFile(suffix=".msg", delete=True) as tmp:
        tmp.write(msg_bytes); tmp.flush()
        m = extract_msg.Message(tmp.name)
        for att in m.attachments:
            fname = getattr(att, "longFilename", None) or getattr(att, "shortFilename", None) or "attachment"
            lower = fname.lower()
            if lower.endswith((".xlsx", ".xlsm", ".xls", ".csv")):
                data = att.data
                mime = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if lower.endswith(".xlsx") else
                    "application/vnd.ms-excel" if lower.endswith((".xls", ".xlsm")) else
                    "text/csv" if lower.endswith(".csv") else
                    mimetypes.guess_type(fname)[0] or "application/octet-stream"
                )
                return fname, data, mime
    return None

def _extract_first_excel_from_eml(eml_bytes: bytes):
    msg = BytesParser(policy=policy.default).parsebytes(eml_bytes)
    for part in msg.walk():
        fname = part.get_filename()
        if not fname:
            continue
        lower = fname.lower()
        if lower.endswith((".xlsx", ".xlsm", ".xls", ".csv")):
            data = part.get_payload(decode=True) or b""
            mime = part.get_content_type() or (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if lower.endswith(".xlsx") else
                "application/vnd.ms-excel" if lower.endswith((".xls", ".xlsm")) else
                "text/csv" if lower.endswith(".csv") else
                mimetypes.guess_type(fname)[0] or "application/octet-stream"
            )
            return fname, data, mime
    return None

def _extract_first_allowed_from_msg(msg_bytes: bytes) -> Optional[Tuple[str, bytes, str, str]]:
    """
    .msg から Excel -> PDF -> Word の優先順で最初の添付を抽出
    戻り値: (filename, data_bytes, mime, kind) / None
    """
    with tempfile.NamedTemporaryFile(suffix=".msg", delete=True) as tmp:
        tmp.write(msg_bytes); tmp.flush()
        m = extract_msg.Message(tmp.name)

        # 1回目スキャンで分類
        candidates = []
        for att in m.attachments:
            fname = getattr(att, "longFilename", None) or getattr(att, "shortFilename", None) or "attachment"
            data = getattr(att, "data", None)
            if not data:
                continue
            guess_mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
            kind = _classify_by_name_mime(fname, guess_mime)
            if kind in ("excel", "pdf", "word"):
                # 正確なMIMEに補正
                if kind == "excel":
                    if fname.lower().endswith(".xlsx"):
                        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif fname.lower().endswith((".xls", ".xlsm")):
                        mime = "application/vnd.ms-excel"
                    elif fname.lower().endswith(".csv"):
                        mime = "text/csv"
                    else:
                        mime = guess_mime
                elif kind == "pdf":
                    mime = "application/pdf"
                elif kind == "word":
                    if fname.lower().endswith(".docx"):
                        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                    else:
                        mime = "application/msword"
                else:
                    mime = guess_mime
                candidates.append((fname, data, mime, kind))

        # 優先順で返す
        for want in ("excel", "pdf", "word"):
            for c in candidates:
                if c[3] == want:
                    return c
    return None

def _extract_first_allowed_from_eml(eml_bytes: bytes) -> Optional[Tuple[str, bytes, str, str]]:
    """
    .eml から Excel -> PDF -> Word の優先順で最初の添付を抽出
    戻り値: (filename, data_bytes, mime, kind) / None
    """
    msg = BytesParser(policy=policy.default).parsebytes(eml_bytes)
    candidates = []
    for part in msg.walk():
        fname = part.get_filename()
        cdisp = part.get_content_disposition()
        ctype = part.get_content_type()
        if cdisp == "attachment" or fname:
            kind = _classify_by_name_mime(fname or "", ctype or "")
            if kind in ("excel", "pdf", "word"):
                data = part.get_payload(decode=True) or b""
                if not data:
                    continue
                # MIME正規化
                if kind == "excel":
                    if (fname or "").lower().endswith(".xlsx"):
                        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif (fname or "").lower().endswith((".xls", ".xlsm")):
                        mime = "application/vnd.ms-excel"
                    elif (fname or "").lower().endswith(".csv"):
                        mime = "text/csv"
                    else:
                        mime = ctype or "application/octet-stream"
                elif kind == "pdf":
                    mime = "application/pdf"
                elif kind == "word":
                    if (fname or "").lower().endswith(".docx"):
                        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                    else:
                        mime = "application/msword"
                else:
                    mime = ctype or "application/octet-stream"
                candidates.append((fname or "attachment", data, mime, kind))

    for want in ("excel", "pdf", "word"):
        for c in candidates:
            if c[3] == want:
                return c
    return None

# ===============================
# OneDrive系 ルーティング
# ===============================
@app.get("/")
def index():
    return jsonify({"ok": True, "service": "onedrive-uploader-org", "version": "2025-10-03"})

@app.get("/front")
def serve_front():
    return send_from_directory(app.template_folder, "onedrive.html")

@app.get("/picker-redirect.html")
def picker_redirect():
    html_s = "<!doctype html><meta charset='utf-8'><title>ok</title>OK"
    resp = app.response_class(html_s, mimetype="text/html")
    resp.headers.pop("Content-Security-Policy", None)
    return resp

# --- 子フォルダ一覧
@app.get("/api/drive/folders")
def api_list_folders():
    parent_id = request.args.get("parentId", "root")
    try:
        folders = graph_list_child_folders(parent_id)
        parent = None if parent_id == "root" else graph_get_item_parent(parent_id)
        return jsonify({"ok": True, "parent": parent, "folders": folders})
    except requests.HTTPError as e:
        return jsonify({"error": "graph_http_error", "status": e.response.status_code, "detail": e.response.text}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# --- 新規フォルダ作成（JSON/Form/Query 互換）
@app.post("/api/drive/create-folder")
@app.post("/api/drive/createFolder")
def api_create_folder():
    try:
        j = request.get_json(silent=True) or {}
        parent_id = (
            request.args.get("parentId")
            or request.form.get("parentId")
            or j.get("parentId")
            or "root"
        )
        name = (
            request.args.get("name")
            or request.form.get("name")
            or j.get("name")
        )
        behavior = (
            request.args.get("conflictBehavior")
            or request.form.get("conflictBehavior")
            or j.get("conflictBehavior")
            or "rename"
        )
        if not name:
            return jsonify({"error": "missing name"}), 400
        item = graph_create_folder(parent_id, name, behavior)
        return jsonify({"ok": True, "id": item.get("id"), "name": item.get("name"), "webUrl": item.get("webUrl")})
    except requests.HTTPError as e:
        return jsonify({"error": "graph_http_error", "status": e.response.status_code, "detail": e.response.text}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# --- フォルダ削除
@app.post("/api/drive/delete-folder")
@app.delete("/api/drive/delete-folder")
def api_delete_folder():
    try:
        j = request.get_json(silent=True) or {}
        item_id = (
            request.args.get("id")
            or request.form.get("id")
            or j.get("id")
        )
        if not item_id:
            return jsonify({"error": "missing id"}), 400
        graph_delete_item(item_id)
        return jsonify({"ok": True})
    except requests.HTTPError as e:
        return jsonify({"error": "graph_http_error", "status": e.response.status_code, "detail": e.response.text}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# --- Tickets: create
@app.post("/tickets/create")
def tickets_create():
    """
    JSONでチケット作成
      {"type":"base64","fileName":"Report.eml","mime":"message/rfc822","data":"<b64>","ttlSec":600}
      {"type":"text","fileName":"note.txt","mime":"text/plain","data":"hello"}
      {"type":"url","fileName":"download.bin","href":"https://..."}
      {"type":"eml","fileName":"Report.eml","payload":{...},"data":"<b64>"}
    """
    try:
        j = request.get_json(force=True, silent=False)
        mtype = (j.get("type") or "text").lower()
        meta = {
            "type": mtype,
            "fileName": j.get("fileName") or ("llm.eml" if mtype == "eml" else "download.bin"),
            "mime": j.get("mime"),
        }
        if mtype in ("text", "base64"):
            meta["data"] = j.get("data") or ""
        elif mtype == "url":
            meta["href"] = j.get("href") or ""
        elif mtype == "eml":
            meta["payload"] = j.get("payload") or {}
            if j.get("data"):
                meta["data"] = j.get("data")
            if not meta.get("mime"):
                meta["mime"] = "message/rfc822"
        else:
            meta["data"] = j.get("data") or ""
        ttl = int(j.get("ttlSec") or DEFAULT_TICKET_TTL)
        tid = save_ticket(meta, ttl=ttl)
        return jsonify({"ticket": tid})
    except Exception as e:
        return jsonify({ "error": str(e)}), 400

# --- Tickets: create-multipart（ファイル＋メタ／msg・eml→添付抽出オプション）
@app.post("/tickets/create-multipart")
def tickets_create_multipart():
    """
    form-data:
      file      : (任意) アップロードファイル（.msg / .eml 等）
      metadata  : (任意) JSON文字列
        例のオプション:
          ttlSec: number
          fileName, mime
          type: "base64"|"text"|"eml"|...
          data: (typeに応じて)
          extractXlsx: true    # 旧仕様: Excel/CSV 抽出
          xlsxFileName: "Attachment.xlsx"
          extractFirstAllowed: true  # 新仕様: Excel→PDF→Word の順で最初の1件を抽出
          allowedHintFileName: "Attachment"  # 保存名ベース（拡張子は添付に合わせる）
    戻り:
      {"ok":true,"tickets":{"file":<ticket>|None,"metadata":<ticket>|None,"xlsxFromMsg":<ticket>|None,"firstAllowed":<ticket>|None,"firstAllowedName": "...", "firstAllowedKind": "excel|pdf|word"}}
    """
    try:
        meta_text = request.form.get("metadata", "") or ""
        if not meta_text and "metadata" in request.files:
            try:
                meta_text = request.files["metadata"].read().decode("utf-8", errors="ignore")
            except Exception:
                meta_text = ""
        if not meta_text:
            meta_text = request.form.get("meta", "") or request.form.get("metadata_json", "") or ""

        meta_json = {}
        if meta_text.strip():
            try:
                meta_json = json.loads(meta_text)
            except Exception:
                meta_json = {}

        ttl = int(meta_json.get("ttlSec") or DEFAULT_TICKET_TTL)

        made_file_ticket = None
        made_meta_ticket = None
        made_xlsx_ticket = None
        made_first_allowed_ticket = None
        made_first_allowed_name = None
        made_first_allowed_kind = None

        f = request.files.get("file")
        raw = None

        if f:
            raw = f.read()
            up_name = getattr(f, "filename", None) or "upload.bin"
            file_name_for_file = (meta_json.get("fileName") or up_name or "upload.bin").strip() or "upload.bin"
            file_mime_for_file = (
                meta_json.get("mime")
                or f.mimetype
                or mimetypes.guess_type(file_name_for_file)[0]
                or "application/octet-stream"
            )
            made_file_ticket = save_ticket({
                "type": "base64",
                "fileName": file_name_for_file,
                "mime": file_mime_for_file,
                "data": base64.b64encode(raw).decode("ascii"),
            }, ttl=ttl)

        if meta_json:
            mtype = (meta_json.get("type") or "text").lower()
            meta = {
                "type": mtype,
                "fileName": meta_json.get("fileName") or "download.bin",
                "mime": meta_json.get("mime"),
            }
            if mtype in ("text", "base64"):
                meta["data"] = meta_json.get("data") or ""
            elif mtype == "url":
                meta["href"] = meta_json.get("href") or ""
            elif mtype == "eml":
                meta["payload"] = meta_json.get("payload") or {}
                if meta_json.get("data"):
                    meta["data"] = meta_json.get("data")
                if not meta.get("mime"):
                    meta["mime"] = "message/rfc822"
            else:
                meta["data"] = meta_json.get("data") or ""
            made_meta_ticket = save_ticket(meta, ttl=ttl)

        # 旧: Excel/CSV 抽出
        if f and raw is not None and meta_json.get("extractXlsx"):
            up_lower = (getattr(f, "filename", "") or "").lower()
            hit = None
            if up_lower.endswith(".msg"):
                hit = _extract_first_excel_from_msg(raw)
            elif up_lower.endswith(".eml"):
                hit = _extract_first_excel_from_eml(raw)
            else:
                hit = _extract_first_excel_from_msg(raw) or _extract_first_excel_from_eml(raw)

            if hit:
                att_name, att_bytes, att_mime = hit
                save_name = (meta_json.get("xlsxFileName") or att_name or "attachment.xlsx").strip()
                if not save_name.lower().endswith((".xlsx", ".xlsm", ".xls", ".csv")):
                    save_name += ".xlsx"
                made_xlsx_ticket = save_ticket({
                    "type": "base64",
                    "fileName": save_name,
                    "mime": att_mime or "application/octet-stream",
                    "data": base64.b64encode(att_bytes).decode("ascii")
                }, ttl=ttl)

        # 新: Excel→PDF→Word の順で最初の1件を抽出
        if f and raw is not None and meta_json.get("extractFirstAllowed"):
            up_lower = (getattr(f, "filename", "") or "").lower()
            hit2 = None
            if up_lower.endswith(".msg"):
                hit2 = _extract_first_allowed_from_msg(raw)
            elif up_lower.endswith(".eml"):
                hit2 = _extract_first_allowed_from_eml(raw)
            else:
                hit2 = _extract_first_allowed_from_msg(raw) or _extract_first_allowed_from_eml(raw)

            if hit2:
                att_name, att_bytes, att_mime, kind = hit2
                # 保存名ベースがあれば使い、拡張子は添付に合わせる
                base = (meta_json.get("allowedHintFileName") or att_name or "attachment").strip()
                # 末尾の拡張子を落として拡張子を付け直す
                root, ext = os.path.splitext(att_name or "")
                if not ext:
                    # 付け直し用推定
                    if kind == "excel":
                        ext = ".xlsx"
                    elif kind == "pdf":
                        ext = ".pdf"
                    elif kind == "word":
                        ext = ".docx"
                    else:
                        ext = ""
                # base から既存拡張子を除去
                base_root, _ = os.path.splitext(base)
                save_name = (base_root or "attachment") + ext
                made_first_allowed_ticket = save_ticket({
                    "type": "base64",
                    "fileName": save_name,
                    "mime": att_mime or "application/octet-stream",
                    "data": base64.b64encode(att_bytes).decode("ascii")
                }, ttl=ttl)
                made_first_allowed_name = save_name
                made_first_allowed_kind = kind

        return jsonify({"ok": True, "tickets": {
            "file": made_file_ticket,
            "metadata": made_meta_ticket,
            "xlsxFromMsg": made_xlsx_ticket,
            "firstAllowed": made_first_allowed_ticket,
            "firstAllowedName": made_first_allowed_name,
            "firstAllowedKind": made_first_allowed_kind
        }})
    except Exception as e:
        print("❌ Error in /tickets/create-multipart:", e)
        traceback.print_exc()  # ← スタックトレースを標準出力へ
        return jsonify({"error": str(e)}), 400

# --- Tickets: peek
@app.get("/tickets/peek")
def tickets_peek():
    t = request.args.get("ticket", "")
    try:
        meta = redeem_ticket(t, consume=False)
        return jsonify({"fileName": meta.get("fileName"), "type": meta.get("type"), "mime": meta.get("mime")})
    except KeyError:
        return jsonify({"error": "ticket_not_found_or_expired"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ---- チケットユーティリティ: .msg/.eml → xlsx チケット化（非消費 & 直ファイル対応）
@app.post("/api/msg-to-xlsx-ticket")
def api_msg_to_xlsx_ticket():
    """
    入力（どれでも可）:
      - form-data / query:
          ticket=<.msg/.eml のチケット> / msg_ticket= / fileName= / ttlSec=
          file=<.msg/.eml を直接アップロード>
      - JSON:
          {"ticket":"<MSG_OR_EML_TICKET>","fileName":"Attachment.xlsx","ttlSec":600}
          {"tickets":{"file":"<MSG_TICKET>"}}
          {"arg1":"{\"ticket\":\"...\",\"fileName\":\"...\"}"}
          {"data":"<base64-raw>","fileName":"mail.eml","ttlSec":600}
    出力: {"ok": true, "ticket": "<XLSX_TICKET>", "fileName": "…"}
    """
    try:
        msg_tid = (
            request.form.get("ticket")
            or request.form.get("msg_ticket")
            or request.args.get("ticket")
            or request.args.get("msg_ticket")
        )
        name_ovr = request.form.get("fileName") or request.args.get("fileName")
        ttl_override = request.form.get("ttlSec") or request.args.get("ttlSec")

        j = request.get_json(silent=True) or {}
        if not msg_tid:
            msg_tid = j.get("ticket") or j.get("msg_ticket")
            if not msg_tid and isinstance(j.get("tickets"), dict):
                msg_tid = j["tickets"].get("file") or j["tickets"].get("msg") or j["tickets"].get("ticket_msg")
            if not msg_tid and isinstance(j.get("arg1"), str):
                try:
                    j2 = json.loads(j["arg1"])
                    msg_tid = j2.get("ticket") or j2.get("msg_ticket")
                    if not name_ovr: name_ovr = j2.get("fileName")
                    if not ttl_override: ttl_override = j2.get("ttlSec")
                except Exception:
                    pass
        if not name_ovr:
            name_ovr = j.get("fileName")
        if not ttl_override:
            ttl_override = j.get("ttlSec")

        ttl = int(ttl_override or DEFAULT_TICKET_TTL)

        raw = None
        src_name = None

        f = request.files.get("file") or request.files.get("msg") or request.files.get("eml")
        if f:
            raw = f.read()
            src_name = getattr(f, "filename", None) or "upload.bin"
        elif msg_tid:
            meta = redeem_ticket(msg_tid, consume=False)  # 非消費
            src_name, raw, _ = materialize_bytes(meta)
        elif j.get("data"):
            raw = base64.b64decode(j["data"])
            src_name = j.get("fileName") or "upload.bin"

        if not raw:
            return jsonify({"error": "missing input (ticket or file or data)"}), 400

        lower = (src_name or "").lower()
        hit = None
        if lower.endswith(".msg"):
            hit = _extract_first_excel_from_msg(raw)
        elif lower.endswith(".eml"):
            hit = _extract_first_excel_from_eml(raw)
        else:
            hit = _extract_first_excel_from_msg(raw) or _extract_first_excel_from_eml(raw)

        if not hit:
            return jsonify({"error": "no_excel_attachment_found"}), 404

        att_name, att_bytes, att_mime = hit
        save_name = (name_ovr or att_name or "attachment.xlsx").strip()
        if not save_name.lower().endswith((".xlsx", ".xlsm", ".xls", ".csv")):
            save_name += ".xlsx"

        x_tid = save_ticket({
            "type": "base64",
            "fileName": save_name,
            "mime": att_mime or "application/octet-stream",
            "data": base64.b64encode(att_bytes).decode("ascii")
        }, ttl=ttl)

        return jsonify({"ok": True, "ticket": x_tid, "fileName": save_name})

    except KeyError:
        return jsonify({"error": "ticket_not_found_or_expired"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ---- NEW: .msg/.eml → Excel/PDF/Word のうち最初の1件をチケット化（非消費 & 直ファイル対応）
@app.post("/api/msg-to-attachment-ticket")
def api_msg_to_attachment_ticket():
    """
    入力: /api/msg-to-xlsx-ticket と同様（ticket / file / data いずれか）
    出力: {"ok": true, "ticket": "<TICKET>", "fileName": "…", "kind": "excel|pdf|word"}
    """
    try:
        msg_tid = (
            request.form.get("ticket")
            or request.form.get("msg_ticket")
            or request.args.get("ticket")
            or request.args.get("msg_ticket")
        )
        name_ovr = request.form.get("fileName") or request.args.get("fileName")
        ttl_override = request.form.get("ttlSec") or request.args.get("ttlSec")

        j = request.get_json(silent=True) or {}
        if not msg_tid:
            msg_tid = j.get("ticket") or j.get("msg_ticket")
            if not msg_tid and isinstance(j.get("tickets"), dict):
                msg_tid = j["tickets"].get("file") or j["tickets"].get("msg") or j["tickets"].get("ticket_msg")
            if not msg_tid and isinstance(j.get("arg1"), str):
                try:
                    j2 = json.loads(j["arg1"])
                    msg_tid = j2.get("ticket") or j2.get("msg_ticket")
                    if not name_ovr: name_ovr = j2.get("fileName")
                    if not ttl_override: ttl_override = j2.get("ttlSec")
                except Exception:
                    pass
        if not name_ovr:
            name_ovr = j.get("fileName")
        if not ttl_override:
            ttl_override = j.get("ttlSec")

        ttl = int(ttl_override or DEFAULT_TICKET_TTL)

        raw = None
        src_name = None

        f = request.files.get("file") or request.files.get("msg") or request.files.get("eml")
        if f:
            raw = f.read()
            src_name = getattr(f, "filename", None) or "upload.bin"
        elif msg_tid:
            meta = redeem_ticket(msg_tid, consume=False)
            src_name, raw, _ = materialize_bytes(meta)
        elif j.get("data"):
            raw = base64.b64decode(j["data"])
            src_name = j.get("fileName") or "upload.bin"

        if not raw:
            return jsonify({"error": "missing input (ticket or file or data)"}), 400

        lower = (src_name or "").lower()
        hit = None
        if lower.endswith(".msg"):
            hit = _extract_first_allowed_from_msg(raw)
        elif lower.endswith(".eml"):
            hit = _extract_first_allowed_from_eml(raw)
        else:
            hit = _extract_first_allowed_from_msg(raw) or _extract_first_allowed_from_eml(raw)

        if not hit:
            return jsonify({"error": "no_allowed_attachment_found"}), 404

        att_name, att_bytes, att_mime, kind = hit

        # 保存名：name_ovr があればそのルート + 添付拡張子、なければ att_name を使用
        if name_ovr:
            root, _ = os.path.splitext(name_ovr.strip())
            ext = os.path.splitext(att_name or "")[1] or (".xlsx" if kind == "excel" else ".pdf" if kind == "pdf" else ".docx")
            save_name = (root or "attachment") + ext
        else:
            save_name = att_name or "attachment"

        t_id = save_ticket({
            "type": "base64",
            "fileName": save_name,
            "mime": att_mime or "application/octet-stream",
            "data": base64.b64encode(att_bytes).decode("ascii")
        }, ttl=ttl)

        return jsonify({"ok": True, "ticket": t_id, "fileName": save_name, "kind": kind})

    except KeyError:
        return jsonify({"error": "ticket_not_found_or_expired"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ---- OneDrive: アイテムメタ取得
@app.get("/api/drive/item")
def api_drive_item():
    item_id = request.args.get("id", "")
    if not item_id:
        return jsonify({"error": "missing id"}), 400
    try:
        meta = graph_get_item_meta(item_id)
        return jsonify({"id": meta.get("id"), "name": meta.get("name"), "webUrl": meta.get("webUrl")})
    except requests.HTTPError as e:
        return jsonify({"error": "graph_http_error", "status": e.response.status_code, "detail": e.response.text}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ---- アップロード（組織 OneDrive）
@app.post("/api/upload")
def api_upload():
    """
    form-data: ticket, folderId, (fileName)
    チケット実体を組織OneDriveのフォルダに保存
    """
    try:
        ticket   = request.form.get("ticket")
        folderId = request.form.get("folderId")
        name_ovr = request.form.get("fileName")
        if not ticket or not folderId:
            return jsonify({"error": "missing parameters"}), 400

        meta = redeem_ticket(ticket, consume=True)
        file_name, data_bytes, mime = materialize_bytes(meta)
        if name_ovr:
            file_name = (name_ovr or "").strip() or file_name

        if len(data_bytes) <= SMALL_MAX_BYTES:
            r = graph_put_small_to_folder_org(folderId, file_name, mime, data_bytes)
        else:
            r = graph_put_chunked_to_folder_org(folderId, file_name, data_bytes)

        if r.status_code in (200, 201):
            j = r.json()
            return jsonify({"ok": True, "id": j.get("id"), "webUrl": j.get("webUrl"), "name": j.get("name")})
        else:
            return jsonify({"error": "graph_upload_failed", "status": r.status_code, "detail": r.text}), r.status_code
    except KeyError:
        return jsonify({"error": "ticket_not_found_or_expired"}), 404
    except requests.HTTPError as e:
        return jsonify({"error": "graph_http_error", "status": e.response.status_code, "detail": e.response.text}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ---- .msg 直アップ（最初のExcel/CSVだけ抽出して保存）
@app.post("/api/upload-msg-xlsx")
def api_upload_msg_xlsx():
    """
    form-data: ticket(必須: .msgのチケット), folderId(必須), (fileName)
    .msg から最初のExcel/CSVを抽出して保存
    """
    try:
        ticket   = request.form.get("ticket")
        folderId = request.form.get("folderId")
        name_ovr = request.form.get("fileName")
        if not ticket or not folderId:
            return jsonify({"error": "missing parameters"}), 400

        meta = redeem_ticket(ticket, consume=True)
        _, msg_bytes, _ = materialize_bytes(meta)
        hit = _extract_first_excel_from_msg(msg_bytes)
        if not hit:
            return jsonify({"error": "no_excel_attachment_found"}), 404

        att_name, att_bytes, att_mime = hit
        save_name = (name_ovr or att_name or "attachment.xlsx").strip()
        if not save_name.lower().endswith((".xlsx", ".xlsm", ".xls", ".csv")):
            save_name += ".xlsx"

        if len(att_bytes) <= SMALL_MAX_BYTES:
            r = graph_put_small_to_folder_org(folderId, save_name, att_mime, att_bytes)
        else:
            r = graph_put_chunked_to_folder_org(folderId, save_name, att_bytes)

        if r.status_code in (200, 201):
            j = r.json()
            return jsonify({"ok": True, "id": j.get("id"), "webUrl": j.get("webUrl"), "name": j.get("name")})
        else:
            return jsonify({"error": "graph_upload_failed", "status": r.status_code, "detail": r.text}), r.status_code
    except KeyError:
        return jsonify({"error": "ticket_not_found_or_expired"}), 404
    except requests.HTTPError as e:
        return jsonify({"error": "graph_http_error", "status": e.response.status_code, "detail": e.response.text}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ===============================
# ======== excel_api 追加分: /extract /extract_mail （仕様そのまま） ========
# ===============================

MAX_ROWS = 200
MAX_COLS = 50
MAX_NONEMPTY = 2000  # 非空セル最大数

def to_str(v) -> str:
    if v is None:
        return ""
    s = str(v)
    return (
        s.replace("_x000D_", " ")
         .replace("\t", " ")
         .replace("\r\n", " ")
         .replace("\n", " ")
         .replace("\r", " ")
         .strip()
    )

def _html_to_text(html_s: str) -> str:
    if not html_s:
        return ""
    s = re.sub(r'(?is)<(script|style).*?>.*?</\1>', '', html_s)
    s = re.sub(r'(?is)<br\s*/?>', '\n', s)
    s = re.sub(r'(?is)</p\s*>', '\n', s)
    s = re.sub(r'(?is)<.*?>', '', s)
    s = html.unescape(s)
    return to_str(s)

def _num_to_col(n: int) -> str:
    s = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s.append(chr(65 + rem))
    return "".join(reversed(s))

def _excel_sparse_from_xlsx_bytes(xlsx_bytes: bytes,
                                  sheet_req: str | None = None,
                                  max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = None
    if sheet_req:
        try:
            idx = int(sheet_req)
            names = wb.sheetnames
            if 0 <= idx < len(names):
                ws = wb[names[idx]]
            elif 1 <= idx <= len(names):
                ws = wb[names[idx - 1]]
        except ValueError:
            if sheet_req in wb.sheetnames:
                ws = wb[sheet_req]
    ws = ws or wb.active

    lines, count = [], 0
    for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            txt = to_str(v)
            if not txt:
                continue
            lines.append(f"{cell.coordinate}\t{txt}")
            count += 1
            if count >= max_nonempty:
                lines.append("# ...truncated...")
                break
        if count >= max_nonempty:
            break
    return "\n".join(lines)

def _excel_sparse_from_xls_bytes(xls_bytes: bytes,
                                 max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
    with tempfile.NamedTemporaryFile(delete=True, suffix=".xls") as tmp:
        tmp.write(xls_bytes)
        tmp.flush()
        book = xlrd.open_workbook(tmp.name)
        sheet = book.sheet_by_index(0)

        lines, count = [], 0
        max_r = min(sheet.nrows, max_rows)
        max_c = min(sheet.ncols, max_cols)
        for r in range(max_r):
            for c in range(max_c):
                v = sheet.cell_value(r, c)
                txt = to_str(v)
                if not txt:
                    continue
                coord = f"{_num_to_col(c+1)}{r+1}"
                lines.append(f"{coord}\t{txt}")
                count += 1
                if count >= max_nonempty:
                    lines.append("# ...truncated...")
                    break
            if count >= max_nonempty:
                break
    return "\n".join(lines)

def _excel_sparse_from_bytes(data: bytes,
                             filename: str | None = None,
                             sheet_req: str | None = None,
                             max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
    name = (filename or "").lower()
    if name.endswith(".xls"):
        return _excel_sparse_from_xls_bytes(data, max_rows, max_cols, max_nonempty)
    return _excel_sparse_from_xlsx_bytes(data, sheet_req, max_rows, max_cols, max_nonempty)

@app.route("/extract", methods=["POST"])
def extract():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "file is required (multipart/form-data)"}), 400

    bom_on = (request.form.get("bom", "true").lower() != "false")
    inline_on = (request.form.get("inline", "true").lower() != "false")
    sheet_req = request.form.get("sheet")

    data = f.read()
    if not data:
        return jsonify({"error": "empty file"}), 400

    try:
        payload = _excel_sparse_from_bytes(data, filename=f.filename, sheet_req=sheet_req)
    except Exception as e:
        return jsonify({"error": f"failed to read workbook: {e}"}), 400

    if bom_on:
        payload = "\ufeff" + payload

    headers = {}
    if not inline_on:
        headers["Content-Disposition"] = 'attachment; filename="extract.tsv"'
    return Response(payload, mimetype="text/plain; charset=utf-8", headers=headers)

@app.route("/extract_mail", methods=["POST"])
def extract_mail():
    up = request.files.get("file")
    if not up:
        return jsonify({"error": "file is required (multipart/form-data)"}), 400

    filename = (up.filename or "").lower()
    data = up.read()

    try:
        if filename.endswith(".msg") or _looks_like_msg(data):
            payload = _handle_msg_bytes(data)
        elif filename.endswith(".eml") or _looks_like_eml(data):
            payload = _handle_eml_bytes(data)
        else:
            try:
                payload = _handle_eml_bytes(data)
            except Exception:
                try:
                    payload = _handle_msg_bytes(data)
                except Exception as e:
                    return jsonify({"error": f"unsupported or unreadable mail file: {e}"}), 400
        return Response(json.dumps(payload, ensure_ascii=False),
                        mimetype="application/json; charset=utf-8")
    except Exception as e:
        return jsonify({"error": f"failed to process mail: {e}"}), 400

def _looks_like_msg(b: bytes) -> bool:
    return len(b) >= 8 and b[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

def _looks_like_eml(b: bytes) -> bool:
    head = b[:512].decode("utf-8", errors="ignore")
    return ("From:" in head or "Subject:" in head) and "\n\n" in head

def _handle_msg_bytes(b: bytes) -> Dict:
    with tempfile.NamedTemporaryFile(delete=True, suffix=".msg") as tmp:
        tmp.write(b)
        tmp.flush()
        msg = extract_msg.Message(tmp.name)

    raw_text = to_str(getattr(msg, "body", "") or "")
    raw_html = getattr(msg, "bodyHTML", "") or ""
    body_text = raw_text or _html_to_text(raw_html)

    excel_results: List[Dict] = []
    for att in msg.attachments:
        name = getattr(att, "longFilename", "") or getattr(att, "shortFilename", "") or "attachment"
        data = getattr(att, "data", None)
        if not data:
            continue
        if _is_excel_filename(name):
            try:
                cells_text = _excel_sparse_from_bytes(data, filename=name)
            except Exception as e:
                cells_text = f"# ERROR: excel parse failed: {e}"
            excel_results.append({"filename": name, "cells": cells_text})
    return {"ok": True, "format": "msg", "body_text": body_text, "excel_attachments": excel_results}

def _handle_eml_bytes(b: bytes) -> Dict:
    msg = BytesParser(policy=policy.default).parsebytes(b)

    body_text = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain" and part.get_content_disposition() in (None, "inline"):
                body_text = to_str(part.get_content())
                if body_text:
                    break
        if not body_text:
            for part in msg.walk():
                if part.get_content_type() == "text/html" and part.get_content_disposition() in (None, "inline"):
                    body_text = _html_to_text(part.get_content())
                    if body_text:
                        break
    else:
        ctype = msg.get_content_type()
        if ctype == "text/plain":
            body_text = to_str(msg.get_content())
        elif ctype == "text/html":
            body_text = _html_to_text(msg.get_content())

    excel_results: List[Dict] = []
    for part in msg.walk():
        fname = part.get_filename()
        cdisp = part.get_content_disposition()
        ctype = part.get_content_type()
        if cdisp == "attachment" or fname:
            if _is_excel_filename(fname) or _is_excel_mime(ctype):
                data = part.get_payload(decode=True) or b""
                if not data:
                    continue
                try:
                    cells_text = _excel_sparse_from_bytes(data, filename=fname or "")
                except Exception as e:
                    cells_text = f"# ERROR: excel parse failed: {e}"
                excel_results.append({"filename": fname or "attachment.xlsx", "cells": cells_text})

    return {"ok": True, "format": "eml", "body_text": body_text, "excel_attachments": excel_results}

# ---- GitHub Raw などのURLから直接ダウンロードして保存（第4のファイル）
@app.post("/api/upload-from-url")
def api_upload_from_url():
    """
    JSON:
      {
        "url": "https://raw.githubusercontent.com/<user>/<repo>/<ref>/<path/to/file>",
        "folderId": "<OneDrive itemId>",
        "fileName": "任意の保存名（省略可。省略時はURL末尾名を使用）"
      }
    成功: {"ok": true, "id": "...", "webUrl": "...", "name": "..."}
    """
    try:
        j = request.get_json(force=True, silent=False)
        url = (j.get("url") or j.get("href") or "").strip()
        folder_id = (j.get("folderId") or "").strip()
        name_ovr = (j.get("fileName") or "").strip()

        if not url or not folder_id:
            return jsonify({"error": "missing url or folderId"}), 400

        r = requests.get(url, timeout=60)
        r.raise_for_status()
        data = r.content
        mime = (r.headers.get("Content-Type") or
                mimetypes.guess_type(url)[0] or
                "application/octet-stream")

        if not name_ovr:
            from urllib.parse import urlparse, unquote
            path = unquote(urlparse(url).path or "")
            base = (path.rsplit("/", 1)[-1] or "download.bin")
            name_ovr = base

        if len(data) <= SMALL_MAX_BYTES:
            up = graph_put_small_to_folder_org(folder_id, name_ovr, mime, data)
        else:
            up = graph_put_chunked_to_folder_org(folder_id, name_ovr, data)

        if up.status_code in (200, 201):
            jj = up.json()
            return jsonify({
                "ok": True,
                "id": jj.get("id"),
                "webUrl": jj.get("webUrl"),
                "name": jj.get("name")
            })
        else:
            return jsonify({
                "error": "graph_upload_failed",
                "status": up.status_code,
                "detail": up.text[:4000]
            }), up.status_code

    except requests.HTTPError as e:
        return jsonify({
            "error": "download_http_error",
            "status": getattr(e.response, "status_code", None),
            "detail": getattr(e.response, "text", "")[:4000]
        }), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 400



# ==========================================================
# Routes added by ChatGPT — clean single-attach capable APIs
# ==========================================================

# 1) Word / PDF / Excel 直POST → チケット化
@app.post("/api/attachment-to-ticket")
def api_attachment_to_ticket():
    """
    multipart/form-data:
      file     : *.doc/*.docx/*.pdf/*.xlsx/*.xlsm/*.xls/*.csv (required)
      fileName : 任意（未指定ならアップロード名）
      ttlSec   : 任意（デフォルト DEFAULT_TICKET_TTL）
    JSONでも可:
      {"data":"<base64>", "fileName":"...", "ttlSec":600}
    """
    try:
        import base64, mimetypes, os
        ttl = int(request.form.get("ttlSec") or request.args.get("ttlSec") or DEFAULT_TICKET_TTL)

        f = request.files.get("file")
        if f:
            raw = f.read()
            name = (request.form.get("fileName") or getattr(f, "filename", None) or "attachment").strip()
        else:
            j = request.get_json(silent=True) or {}
            if not j:
                return jsonify({"error": "file or JSON data required"}), 400
            raw = base64.b64decode(j.get("data") or b"")
            name = (j.get("fileName") or "attachment").strip()

        if not raw:
            return jsonify({"error": "empty file"}), 400

        allowed = {
            ".doc":  "application/msword",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".pdf":  "application/pdf",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xlsm": "application/vnd.ms-excel",
            ".xls":  "application/vnd.ms-excel",
            ".csv":  "text/csv",
        }
        ext = (os.path.splitext(name)[1] or "").lower()
        if ext not in allowed:
            return jsonify({"error": "unsupported extension", "fileName": name}), 400
        mime = allowed.get(ext) or mimetypes.guess_type(name)[0] or "application/octet-stream"

        tid = save_ticket({
            "type": "base64",
            "fileName": name,
            "mime": mime,
            "data": base64.b64encode(raw).decode("ascii")
        }, ttl=ttl)

        if ext in (".xlsx", ".xlsm", ".xls", ".csv"):
            kind = "excel"
        elif ext == ".pdf":
            kind = "pdf"
        else:
            kind = "word"

        return jsonify({"ok": True, "ticket": tid, "fileName": name, "kind": kind, "mime": mime})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# 2) Excel 単体 → TSV（/extract_mail と同じ JSON 形）
@app.post("/extract_excel_mailstyle")
def extract_excel_mailstyle():
    """
    multipart/form-data:
      file : *.xlsx/*.xlsm/*.xls/*.csv (required)
      sheet: 任意（名前 or 0/1始まりIndex）
    JSONでも可:
      {"data":"<base64>", "fileName":"...", "sheet":"Sheet1"}

    戻り値例:
      {"ok":true,"format":"excel","excel_attachments":[{"filename":"...","cells":"TSV..."}]}
    """
    try:
        import base64 as _b64, json as _json
        from flask import Response as _Response

        f = request.files.get("file")
        sheet_req = request.form.get("sheet") or request.args.get("sheet")
        if f:
            data = f.read()
            filename = getattr(f, "filename", None) or "attachment.xlsx"
        else:
            j = request.get_json(silent=True) or {}
            if not j:
                return jsonify({"error": "file or JSON data required"}), 400
            data = _b64.b64decode(j.get("data") or b"")
            filename = j.get("fileName") or "attachment.xlsx"
            sheet_req = j.get("sheet") or sheet_req

        if not data:
            return jsonify({"error": "empty file"}), 400

        cells_tsv = _excel_sparse_from_bytes(data, filename=filename, sheet_req=sheet_req)
        payload = {
            "ok": True,
            "format": "excel",
            "excel_attachments": [{
                "filename": filename,
                "cells": "\ufeff" + cells_tsv,  # UTF-8 BOM付与で互換
            }]
        }
        return _Response(_json.dumps(payload, ensure_ascii=False), mimetype="application/json; charset=utf-8")
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# 3) .msg/.eml → (Excel/PDF/Word) 添付をチケット化（複数）
#    単一化オプション: limit=1 / pick=first|0|<index>
@app.post("/api/msg-to-attachment-tickets")
def api_msg_to_attachment_tickets():
    """
    入力: msg_ticket/ticket（チケットID） or multipart 'file' or JSON {data,fileName}
    任意:
      kinds="excel,pdf,word"（フィルタ。省略時は全部）
      baseName="保存名ベース"（拡張子は添付に合わせる）
      ttlSec=秒
      limit=数（例:1）
      pick="first" or "0" or "1"...（index優先指定）

    返却: {"ok":true,"tickets":[{"ticket","fileName","kind","mime"}],"count":N}
    """
    try:
        import base64, os, mimetypes, tempfile
        from email import policy
        from email.parser import BytesParser

        j = request.get_json(silent=True) or {}

        # 入力チケット/ファイル/JSONのいずれか
        msg_tid = (
            request.form.get("ticket") or request.form.get("msg_ticket") or
            request.args.get("ticket") or request.args.get("msg_ticket") or
            j.get("ticket") or j.get("msg_ticket") or (j.get("tickets", {}) or {}).get("file")
        )
        ttl = int(request.form.get("ttlSec") or request.args.get("ttlSec") or j.get("ttlSec") or DEFAULT_TICKET_TTL)
        base_name = (request.form.get("baseName") or request.args.get("baseName") or j.get("baseName") or "").strip()
        kinds_str = (request.form.get("kinds") or request.args.get("kinds") or j.get("kinds") or "").strip().lower()
        kinds_sel = set([s.strip() for s in kinds_str.split(",") if s.strip()]) if kinds_str else {"excel", "pdf", "word"}

        # 単一化オプション
        limit_str = (request.form.get("limit") or request.args.get("limit") or j.get("limit"))
        try:
            limit_n = int(limit_str) if limit_str not in (None, "") else None
        except Exception:
            limit_n = None
        pick = (request.form.get("pick") or request.args.get("pick") or j.get("pick") or "").strip().lower()

        # 元データ取得
        raw = None
        src_name = None
        f = request.files.get("file") or request.files.get("msg") or request.files.get("eml")
        if f:
            raw = f.read()
            src_name = getattr(f, "filename", None) or "upload.bin"
        elif msg_tid:
            meta = redeem_ticket(msg_tid, consume=False)
            src_name, raw, _ = materialize_bytes(meta)
        elif j.get("data"):
            raw = base64.b64decode(j["data"])
            src_name = j.get("fileName") or "upload.bin"

        if not raw:
            return jsonify({"error": "missing input (ticket or file or data)"}), 400

        # 種別判定
        def classify(fname: str, mime: str | None) -> str | None:
            ln = (fname or "").lower()
            if ln.endswith((".xlsx", ".xlsm", ".xls", ".csv")):
                return "excel"
            if ln.endswith(".pdf") or (mime == "application/pdf"):
                return "pdf"
            if ln.endswith((".doc", ".docx")):
                return "word"
            if (mime or "").startswith("application/vnd.openxmlformats-officedocument.spreadsheetml"):
                return "excel"
            if (mime or "").startswith("application/msword") or (mime or "").endswith("wordprocessingml.document"):
                return "word"
            return None

        # 添付抽出
        extracted: list[tuple[str, bytes, str, str]] = []  # (fname, bytes, mime, kind)
        lower = (src_name or "").lower()

        if lower.endswith(".msg"):
            try:
                import extract_msg
            except Exception as e:
                return jsonify({"error": "extract_msg_not_available", "detail": str(e)}), 500

            with tempfile.NamedTemporaryFile(suffix=".msg", delete=True) as tmp:
                tmp.write(raw)
                tmp.flush()
                m = extract_msg.Message(tmp.name)
                for att in getattr(m, "attachments", []) or []:
                    fname = getattr(att, "longFilename", None) or getattr(att, "shortFilename", None) or "attachment"
                    data = getattr(att, "data", None)
                    mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
                    k = classify(fname, mime)
                    if k and k in kinds_sel and data:
                        ln = (fname or "").lower()
                        if k == "excel":
                            if ln.endswith(".xlsx"):
                                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            elif ln.endswith((".xls", ".xlsm")):
                                mime = "application/vnd.ms-excel"
                            elif ln.endswith(".csv"):
                                mime = "text/csv"
                        elif k == "pdf":
                            mime = "application/pdf"
                        else:
                            mime = (
                                "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                                if ln.endswith(".docx") else "application/msword"
                            )
                        extracted.append((fname, data, mime, k))
        else:
            # EML として扱う
            msg = BytesParser(policy=policy.default).parsebytes(raw)
            for part in msg.walk():
                fname = part.get_filename()
                cdisp = part.get_content_disposition()
                ctype = part.get_content_type()
                if cdisp == "attachment" or fname:
                    data = part.get_payload(decode=True) or b""
                    k = classify(fname or "attachment", ctype or None)
                    if k and k in kinds_sel and data:
                        extracted.append((fname or "attachment", data, ctype or "application/octet-stream", k))

        if not extracted:
            return jsonify({"error": "no_allowed_attachment_found", "allowed": sorted(list(kinds_sel))}), 404

        # 単一化（pick/limit）
        if pick in ("first", "0") and extracted:
            extracted = [extracted[0]]
        elif pick.isdigit():
            idx = int(pick)
            if 0 <= idx < len(extracted):
                extracted = [extracted[idx]]

        if limit_n is not None and limit_n >= 0:
            extracted = extracted[:max(1, limit_n)] if extracted else extracted

        # チケット化
        results = []
        for (att_name, att_bytes, att_mime, kind) in extracted:
            if base_name:
                base_root, _ = os.path.splitext(base_name)
                _, ext = os.path.splitext(att_name or "")
                if not ext:
                    ext = ".xlsx" if kind == "excel" else ".pdf" if kind == "pdf" else ".docx"
                save_name = (base_root or "attachment") + ext
            else:
                save_name = att_name or "attachment"

            tid = save_ticket({
                "type": "base64",
                "fileName": save_name,
                "mime": att_mime or "application/octet-stream",
                "data": base64.b64encode(att_bytes).decode("ascii")
            }, ttl=ttl)

            results.append({"ticket": tid, "fileName": save_name, "kind": kind, "mime": att_mime})

        return jsonify({"ok": True, "tickets": results, "count": len(results)})
    except KeyError:
        return jsonify({"error": "ticket_not_found_or_expired"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ===============================
# メイン
# ===============================
# ========= PDF → table text extractor (PyMuPDF) =========
# Requirements:
#   pip install pymupdf
# Usage:
#   1) app.py で Flask を使っている前提（from flask import Flask などで app が存在）
#   2) このブロックを app.py に追記
#   3) POST /extract_pdf_tables に multipart/form-data で file=PDF を投げる
#
# 出力は /extract_mail と同じノリの JSON（※日本語はエスケープしないUTF-8で返却）:
# {
#   "ok": true,
#   "filename": "xxx.pdf",
#   "pages": 2,
#   "rows": 340,
#   "tables_text": "項番 | 在籍期間 | ...\n1 | 2009/7～2011/8 | ...\n...",
#   "suggested_columns": [[...], ...]  # 今は未使用/将来拡張用
# }

from flask import request, jsonify, Response
import json

def _median(vals):
    try:
        import statistics
        return statistics.median(vals) if vals else None
    except Exception:
        return None

def _extract_pdf_to_tabletext_bytes(pdf_bytes: bytes) -> dict:
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        return {"ok": False, "error": f"pymupdf_not_available: {e}"}

    import io
    doc = fitz.open(stream=io.BytesIO(pdf_bytes).getvalue(), filetype="pdf")

    all_lines = []          # 最終的に LLM に渡す "表テキスト" の行
    page_columns = []       # 将来用（列ルーラーのヒント）
    total_rows = 0

    for pno in range(len(doc)):
        page = doc.load_page(pno)

        # 1) span（文字列＋bbox）を取得
        text_dict = page.get_text("dict")
        spans = []
        for b in text_dict.get("blocks", []):
            for l in b.get("lines", []):
                for s in l.get("spans", []):
                    t = (s.get("text") or "").strip()
                    if not t:
                        continue
                    x0, y0, x1, y1 = s.get("bbox", [None, None, None, None])
                    if None in (x0, y0, x1, y1):
                        continue
                    cx = (x0 + x1) / 2.0
                    cy = (y0 + y1) / 2.0
                    spans.append({
                        "text": t,
                        "bbox": [x0, y0, x1, y1],
                        "center": [cx, cy],
                        "size": s.get("size") or 10.0,
                    })

        # 2) 行クラスタ（Y近接）
        spans_sorted = sorted(spans, key=lambda s: (s["center"][1], s["center"][0]))
        sizes = [s["size"] for s in spans_sorted]
        tol_y = (_median(sizes) or 10.0) * 0.6  # 文字サイズベースで動的しきい値

        rows = []
        current_row, last_y = [], None
        for s in spans_sorted:
            cy = s["center"][1]
            if last_y is None or abs(cy - last_y) <= tol_y:
                current_row.append(s)
                last_y = cy if last_y is None else (last_y * 0.6 + cy * 0.4)  # 平滑化
            else:
                rows.append(current_row)
                current_row = [s]
                last_y = cy
        if current_row:
            rows.append(current_row)

        # 3) 行内：X順に並べ、近接 span を結合してセル化
        out_rows = []
        for r in rows:
            r_sorted = sorted(r, key=lambda s: s["center"][0])
            merged, last = [], None
            for s in r_sorted:
                x0, y0, x1, y1 = s["bbox"]
                if last is None:
                    last = {"x0": x0, "x1": x1, "y0": y0, "y1": y1, "text": s["text"], "size": s["size"]}
                else:
                    gap = x0 - last["x1"]
                    fsz = s["size"]
                    if gap < (fsz * 0.6):  # 近い→結合
                        last["x1"] = max(last["x1"], x1)
                        last["y0"] = min(last["y0"], y0)
                        last["y1"] = max(last["y1"], y1)
                        sep = "" if last["text"].endswith(("(", "/", "-", "・")) else " "
                        last["text"] += ("" if s["text"].startswith((")", "/", "-", ",", "・")) else sep) + s["text"]
                    else:
                        merged.append(last)
                        last = {"x0": x0, "x1": x1, "y0": y0, "y1": y1, "text": s["text"], "size": s["size"]}
            if last:
                merged.append(last)
            out_rows.append(merged)

        # 4) 行を " | " 区切りで連結（列ルーラーの厳密化は将来対応）
        for r in out_rows:
            r_sorted = sorted(r, key=lambda c: (c["x0"] + c["x1"]) / 2.0)
            line = " | ".join(c["text"] for c in r_sorted).strip()
            if line:
                all_lines.append(line)

        total_rows += len(out_rows)
        page_columns.append([])  # いったん空（拡張用）

    joined = "\n".join(all_lines)
    return {
        "ok": True,
        "pages": len(doc),
        "rows": total_rows,
        "tables_text": joined,
        "suggested_columns": page_columns,
    }

# ---- Flask endpoint (JSON, Dify-friendly, 日本語エスケープしない) ----
@app.post("/extract_pdf_tables")
def extract_pdf_tables():
    """
    multipart/form-data:
      file: (required) PDF
    return: JSON（/extract_mail を参考にした素朴な形, UTF-8, ensure_ascii=False）
    """
    up = request.files.get("file")
    if not up:
        payload = {"ok": False, "error": "file is required (multipart/form-data)"}
        return Response(json.dumps(payload, ensure_ascii=False), mimetype="application/json; charset=utf-8", status=400)

    fname = up.filename or "upload.pdf"
    data = up.read() or b""
    if not data:
        payload = {"ok": False, "error": "empty file", "filename": fname}
        return Response(json.dumps(payload, ensure_ascii=False), mimetype="application/json; charset=utf-8", status=400)

    try:
        result = _extract_pdf_to_tabletext_bytes(data)
        result["filename"] = fname
        status = 200 if result.get("ok") else 500
        return Response(json.dumps(result, ensure_ascii=False), mimetype="application/json; charset=utf-8", status=status)
    except Exception as e:
        payload = {"ok": False, "error": f"extract_failed: {e}", "filename": fname}
        return Response(json.dumps(payload, ensure_ascii=False), mimetype="application/json; charset=utf-8", status=400)


# ===============================
# ========== EML 編集API (fixed) =
# ===============================
from flask import request, jsonify
import base64, mimetypes, tempfile, os, traceback
from email import policy
from email.parser import BytesParser
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
import html as _html
import re

# 改行入りヘッダーを安全に正規化
def _clean_hdr(v: str) -> str:
    try:
        return re.sub(r'[\r\n]+', ' ', (v or '')).strip()
    except Exception:
        return (v or '').replace('\r',' ').replace('\n',' ').strip()

def _preview_from_eml_bytes(b: bytes) -> dict:
    # 壊れたヘッダーに寛容
    msg = BytesParser(policy=policy.compat32).parsebytes(b)
    subject = _clean_hdr(msg.get('Subject', '') or '')
    from_   = _clean_hdr(msg.get('From', '') or '')
    to_     = _clean_hdr(msg.get('To', '') or '')
    cc_     = _clean_hdr(msg.get('Cc', '') or '')
    bcc_    = _clean_hdr(msg.get('Bcc', '') or '')
    date_   = _clean_hdr(msg.get('Date', '') or '')

    body_text = ''
    body_html = ''

    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain' and part.get_content_disposition() in (None, 'inline'):
                try:
                    body_text = part.get_content()
                except Exception:
                    body_text = (part.get_payload(decode=True) or b'').decode('utf-8', 'ignore')
                break
        for part in msg.walk():
            if part.get_content_type() == 'text/html' and part.get_content_disposition() in (None, 'inline'):
                try:
                    body_html = part.get_content()
                except Exception:
                    body_html = (part.get_payload(decode=True) or b'').decode('utf-8', 'ignore')
                break
    else:
        ctype = msg.get_content_type()
        try:
            content = msg.get_content()
        except Exception:
            content = (msg.get_payload(decode=True) or b'').decode('utf-8', 'ignore')
        if ctype == 'text/html':
            body_html = content or ''
        elif ctype == 'text/plain':
            body_text = content or ''

    atts = []
    for part in msg.walk():
        cd = part.get_content_disposition()
        if cd in ('attachment', 'inline') and part.get_filename():
            atts.append({
                'fileName': part.get_filename(),
                'mime': part.get_content_type(),
                'disposition': cd,
                'size': len(part.get_payload(decode=True) or b'')
            })

    return {
        'subject': subject,
        'body_text': body_text,
        'body_html': body_html,
        'headers': {'from': from_, 'to': to_, 'cc': cc_, 'bcc': bcc_, 'date': date_},
        'attachments': atts
    }

def _preview_from_msg_bytes(b: bytes) -> dict:
    import extract_msg
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, 'mail.msg')
        with open(p, 'wb') as f:
            f.write(b)
        m = extract_msg.Message(p)
        subject = _clean_hdr(getattr(m, 'subject', '') or '')
        body_text = (getattr(m, 'body', '') or '')
        body_html = (getattr(m, 'htmlBody', '') or '')
        from_ = _clean_hdr(getattr(m, 'sender', '') or '')
        to_   = _clean_hdr(getattr(m, 'to', '') or '')
        cc_   = _clean_hdr(getattr(m, 'cc', '') or '')
        bcc_  = _clean_hdr(getattr(m, 'bcc', '') or '')
        date_ = _clean_hdr(getattr(m, 'date', '') or '')
        atts = []
        for a in (getattr(m, 'attachments', None) or []):
            name = getattr(a, 'longFilename', '') or getattr(a, 'shortFilename', '') or 'attachment'
            data = getattr(a, 'data', None)
            if not data:
                continue
            atts.append({'fileName': name, 'mime': mimetypes.guess_type(name)[0] or 'application/octet-stream', 'size': len(data)})
        return {
            'subject': subject,
            'body_text': body_text,
            'body_html': body_html,
            'headers': {'from': from_, 'to': to_, 'cc': cc_, 'bcc': bcc_, 'date': date_},
            'attachments': atts
        }

@app.get('/api/mail/preview-from-ticket')
def api_mail_preview_from_ticket():
    try:
        ticket = request.args.get('ticket')
        if not ticket:
            return jsonify({'ok': False, 'error': 'missing ticket'}), 200
        meta = redeem_ticket(ticket, consume=False)
        file_name, b, mime = materialize_bytes(meta)
        mime = (mime or '').lower()
        if (file_name or '').lower().endswith('.msg') or 'application/vnd.ms-outlook' in mime:
            out = _preview_from_msg_bytes(b)
        else:
            out = _preview_from_eml_bytes(b)
        return jsonify({'ok': True, **out, 'sourceFileName': file_name})
    except KeyError as e:
        return jsonify({'ok': False, 'error': str(e)}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 200

@app.post('/api/mail/compose-from-ticket')
def api_mail_compose_from_ticket():
    try:
        j = request.get_json(force=True, silent=False)
        ticket = j.get('ticket')
        if not ticket:
            return jsonify({'error': 'missing ticket'}), 400
        subject = j.get('subject') or ''
        body_html = j.get('body_html') or ''
        body_text = j.get('body_text') or ''
        keep_atts = bool(j.get('keep_attachments', True))
        regen     = bool(j.get('regenerate_date', True))
        base_name = (j.get('suggestedBaseName') or '').strip()

        if not (body_html or body_text):
            return jsonify({'error': 'body_html or body_text is required'}), 400

        meta = redeem_ticket(ticket, consume=False)
        orig_name, orig_bytes, orig_mime = materialize_bytes(meta)

        from_addr = ''
        to_addr = ''
        cc_addr = ''
        bcc_addr = ''
        date_hdr = ''
        orig_atts = []

        try:
            if (orig_name or '').lower().endswith('.msg') or 'application/vnd.ms-outlook' in (orig_mime or '').lower():
                import extract_msg
                with tempfile.TemporaryDirectory() as td:
                    p = os.path.join(td, 'mail.msg')
                    with open(p, 'wb') as f:
                        f.write(orig_bytes)
                    m = extract_msg.Message(p)
                    from_addr = _clean_hdr(getattr(m, 'sender', '') or '')
                    to_addr   = _clean_hdr(getattr(m, 'to', '') or '')
                    cc_addr   = _clean_hdr(getattr(m, 'cc', '') or '')
                    bcc_addr  = _clean_hdr(getattr(m, 'bcc', '') or '')
                    date_hdr  = _clean_hdr(getattr(m, 'date', '') or '')
                    if keep_atts:
                        for a in (getattr(m, 'attachments', None) or []):
                            name = getattr(a, 'longFilename', '') or getattr(a, 'shortFilename', '') or 'attachment'
                            data = getattr(a, 'data', None)
                            if not data:
                                continue
                            mime = mimetypes.guess_type(name)[0] or 'application/octet-stream'
                            orig_atts.append((name, mime, data))
            else:
                em = BytesParser(policy=policy.compat32).parsebytes(orig_bytes)
                from_addr = _clean_hdr(em.get('From', '') or '')
                to_addr   = _clean_hdr(em.get('To', '') or '')
                cc_addr   = _clean_hdr(em.get('Cc', '') or '')
                bcc_addr  = _clean_hdr(em.get('Bcc', '') or '')
                date_hdr  = _clean_hdr(em.get('Date', '') or '')
                if keep_atts:
                    for part in em.walk():
                        cd = part.get_content_disposition()
                        fn = part.get_filename()
                        if cd in ('attachment', 'inline') and fn:
                            data = part.get_payload(decode=True) or b''
                            mime = part.get_content_type() or 'application/octet-stream'
                            orig_atts.append((fn, mime, data))
        except Exception:
            pass

        new = EmailMessage()
        if subject:
            new['Subject'] = _clean_hdr(subject)
        if from_addr:
            new['From'] = _clean_hdr(from_addr)
        if to_addr:
            new['To'] = _clean_hdr(to_addr)
        if cc_addr:
            new['Cc'] = _clean_hdr(cc_addr)
        if bcc_addr:
            new['Bcc'] = _clean_hdr(bcc_addr)

        if regen or not date_hdr:
            new['Date'] = formatdate(localtime=True)
            new['Message-ID'] = make_msgid()
        else:
            new['Date'] = _clean_hdr(date_hdr)
            new['Message-ID'] = make_msgid()

        if body_text:
            new.set_content(body_text)
            if body_html:
                new.add_alternative(body_html, subtype='html')
        else:
            import re as _re2
            new.set_content(_html.unescape(_re2.sub(r'<[^>]+>', '', body_html)))
            new.add_alternative(body_html, subtype='html')

        for fn, mime, data in orig_atts:
            maintype, subtype = (mime.split('/', 1) + ['octet-stream'])[:2]
            new.add_attachment(data, maintype=maintype, subtype=subtype, filename=fn)

        out_bytes = new.as_bytes()
        if not base_name:
            base, _ = os.path.splitext(orig_name or 'mail')
            base_name = f"{base}_edited"
        suggested = base_name + '.eml'
        meta2 = {
            'type': 'base64',
            'fileName': suggested,
            'mime': 'message/rfc822',
            'data': base64.b64encode(out_bytes).decode('ascii')
        }
        tid2 = save_ticket(meta2, ttl=DEFAULT_TICKET_TTL)
        return jsonify({'ok': True, 'ticket_eml': tid2, 'suggestedFileName': suggested, 'size': len(out_bytes)})
    except KeyError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 400

